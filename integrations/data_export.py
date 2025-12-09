#!/usr/bin/env python3
"""
Data Export - Export members, campaigns, and analytics to various formats
"""

import csv
import json
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from PyQt6.QtCore import pyqtSignal
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QButtonGroup,
    QCheckBox,
    QDialog,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QVBoxLayout,
)

from integrations.export_manager import normalize_export_path

logger = logging.getLogger(__name__)


class DataExportDialog(QDialog):
    """Dialog for exporting data."""

    export_started = pyqtSignal(dict)  # Export configuration

    def __init__(self, data_type: str = "members", data_count: int = 0, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"📤 Export {data_type.title()}")
        self.resize(600, 500)

        self.data_type = data_type
        self.data_count = data_count
        self.export_config = {}

        self.setup_ui()

    def setup_ui(self):
        """Setup the UI."""
        layout = QVBoxLayout(self)

        # Title
        title = QLabel(f"📤 Export {self.data_type.title()}")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        desc = QLabel(
            f"Export {self.data_count:,} {self.data_type} to file.\n"
            "Choose format and options below."
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)

        # Format selection
        format_group = QGroupBox("📁 Export Format")
        format_layout = QVBoxLayout()

        self.format_group = QButtonGroup(self)

        self.csv_radio = QRadioButton("CSV (Comma-Separated Values)")
        self.csv_radio.setToolTip("Best for Excel, Google Sheets, and data analysis")
        self.csv_radio.setChecked(True)
        self.format_group.addButton(self.csv_radio, 0)
        format_layout.addWidget(self.csv_radio)

        self.json_radio = QRadioButton("JSON (JavaScript Object Notation)")
        self.json_radio.setToolTip("Best for developers and API integration")
        self.format_group.addButton(self.json_radio, 1)
        format_layout.addWidget(self.json_radio)

        self.excel_radio = QRadioButton("Excel (.xlsx)")
        self.excel_radio.setToolTip("Native Excel format with formatting")
        self.excel_radio.setEnabled(False)  # Needs openpyxl
        self.format_group.addButton(self.excel_radio, 2)
        format_layout.addWidget(self.excel_radio)

        try:
            import openpyxl  # noqa: F401

            self.excel_radio.setEnabled(True)
        except ImportError:
            self.excel_radio.setToolTip("Install openpyxl to enable: pip install openpyxl")

        self.txt_radio = QRadioButton("Plain Text (.txt)")
        self.txt_radio.setToolTip("Simple text file, one entry per line")
        self.format_group.addButton(self.txt_radio, 3)
        format_layout.addWidget(self.txt_radio)

        format_group.setLayout(format_layout)
        layout.addWidget(format_group)

        # Options based on data type
        if self.data_type == "members":
            self.setup_member_options(layout)
        elif self.data_type == "campaigns":
            self.setup_campaign_options(layout)
        elif self.data_type == "accounts":
            self.setup_account_options(layout)

        # File location
        location_group = QGroupBox("💾 Save Location")
        location_layout = QHBoxLayout()

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("Choose save location...")
        self.file_path_edit.setReadOnly(True)
        location_layout.addWidget(self.file_path_edit)

        browse_btn = QPushButton("📁 Browse...")
        browse_btn.clicked.connect(self.browse_location)
        location_layout.addWidget(browse_btn)

        location_group.setLayout(location_layout)
        layout.addWidget(location_group)

        # Status
        self.status_label = QLabel("")
        self.status_label.setStyleSheet(
            "padding: 10px; background-color: #2b2d31; border-radius: 4px;"
        )
        layout.addWidget(self.status_label)

        # Buttons
        button_layout = QHBoxLayout()

        self.export_btn = QPushButton("📤 Export")
        self.export_btn.clicked.connect(self.start_export)
        self.export_btn.setObjectName("success")
        button_layout.addWidget(self.export_btn)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        button_layout.addStretch()
        layout.addLayout(button_layout)

        self.update_status()

    def setup_member_options(self, parent_layout):
        """Setup member-specific export options."""
        options_group = QGroupBox("⚙️ Export Options")
        options_layout = QVBoxLayout()

        self.include_phone_check = QCheckBox("Include phone numbers (if available)")
        self.include_phone_check.setToolTip(
            "Phone numbers are rarely available for scraped members"
        )
        options_layout.addWidget(self.include_phone_check)

        self.include_bio_check = QCheckBox("Include bio/about")
        self.include_bio_check.setChecked(True)
        options_layout.addWidget(self.include_bio_check)

        self.include_status_check = QCheckBox("Include online status")
        self.include_status_check.setChecked(True)
        options_layout.addWidget(self.include_status_check)

        self.include_dates_check = QCheckBox("Include join/scrape dates")
        self.include_dates_check.setChecked(True)
        options_layout.addWidget(self.include_dates_check)

        self.anonymize_check = QCheckBox("Anonymize data (remove IDs)")
        self.anonymize_check.setToolTip("Remove Telegram IDs for privacy")
        options_layout.addWidget(self.anonymize_check)

        options_group.setLayout(options_layout)
        parent_layout.addWidget(options_group)

    def setup_campaign_options(self, parent_layout):
        """Setup campaign-specific export options."""
        options_group = QGroupBox("⚙️ Export Options")
        options_layout = QVBoxLayout()

        self.include_messages_check = QCheckBox("Include message templates")
        self.include_messages_check.setChecked(True)
        options_layout.addWidget(self.include_messages_check)

        self.include_stats_check = QCheckBox("Include detailed statistics")
        self.include_stats_check.setChecked(True)
        options_layout.addWidget(self.include_stats_check)

        self.include_timeline_check = QCheckBox("Include timeline/history")
        options_layout.addWidget(self.include_timeline_check)

        options_group.setLayout(options_layout)
        parent_layout.addWidget(options_group)

    def setup_account_options(self, parent_layout):
        """Setup account-specific export options."""
        options_group = QGroupBox("⚙️ Export Options")
        options_layout = QVBoxLayout()

        self.include_sessions_check = QCheckBox("Include session info")
        self.include_sessions_check.setChecked(True)
        options_layout.addWidget(self.include_sessions_check)

        self.include_stats_check = QCheckBox("Include usage statistics")
        self.include_stats_check.setChecked(True)
        options_layout.addWidget(self.include_stats_check)

        self.include_proxy_check = QCheckBox("Include proxy configuration")
        options_layout.addWidget(self.include_proxy_check)

        options_group.setLayout(options_layout)
        parent_layout.addWidget(options_group)

    def browse_location(self):
        """Browse for save location."""
        format_id = self.format_group.checkedId()

        if format_id == 0:  # CSV
            filter_str = "CSV Files (*.csv)"
            default_ext = "csv"
        elif format_id == 1:  # JSON
            filter_str = "JSON Files (*.json)"
            default_ext = "json"
        elif format_id == 2:  # Excel
            filter_str = "Excel Files (*.xlsx)"
            default_ext = "xlsx"
        else:  # TXT
            filter_str = "Text Files (*.txt)"
            default_ext = "txt"

        # Default filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"{self.data_type}_export_{timestamp}.{default_ext}"

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Export File", default_name, filter_str
        )

        if file_path:
            self.file_path_edit.setText(file_path)
            self.update_status()

    def update_status(self):
        """Update status display."""
        if self.file_path_edit.text():
            self.status_label.setText(
                f"✅ Ready to export {self.data_count:,} {self.data_type} to:\n"
                f"{self.file_path_edit.text()}"
            )
            self.export_btn.setEnabled(True)
        else:
            self.status_label.setText("⚠️ Choose a save location to continue")
            self.export_btn.setEnabled(False)

    def get_export_config(self) -> Dict[str, Any]:
        """Get export configuration."""
        config = {
            "data_type": self.data_type,
            "file_path": self.file_path_edit.text(),
            "format": ["csv", "json", "xlsx", "txt"][self.format_group.checkedId()],
        }

        # Add type-specific options
        if self.data_type == "members":
            config["options"] = {
                "include_phone": self.include_phone_check.isChecked(),
                "include_bio": self.include_bio_check.isChecked(),
                "include_status": self.include_status_check.isChecked(),
                "include_dates": self.include_dates_check.isChecked(),
                "anonymize": self.anonymize_check.isChecked(),
            }
        elif self.data_type == "campaigns":
            config["options"] = {
                "include_messages": self.include_messages_check.isChecked(),
                "include_stats": self.include_stats_check.isChecked(),
                "include_timeline": self.include_timeline_check.isChecked(),
            }
        elif self.data_type == "accounts":
            config["options"] = {
                "include_sessions": self.include_sessions_check.isChecked(),
                "include_stats": self.include_stats_check.isChecked(),
                "include_proxy": self.include_proxy_check.isChecked(),
            }

        return config

    def start_export(self):
        """Start the export process."""
        if not self.file_path_edit.text():
            QMessageBox.warning(self, "No Location", "Please choose a save location")
            return

        self.export_config = self.get_export_config()
        self.export_started.emit(self.export_config)
        self.accept()

    def get_config(self) -> Dict[str, Any]:
        """Get the export configuration."""
        return self.export_config


class DataExporter:
    """Handles actual data export operations with REAL database integration."""

    @staticmethod
    def fetch_real_member_data(filter_criteria: Dict = None) -> List[Dict]:
        """Fetch ACTUAL member data from database."""
        import sqlite3

        from member_scraper import MemberDatabase

        try:
            member_db = MemberDatabase("members.db")
            conn = sqlite3.connect(member_db.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            # Build WHERE clause from filter criteria
            conditions = []
            params = []

            if filter_criteria:
                if filter_criteria.get("has_username"):
                    conditions.append("username IS NOT NULL AND username != ''")

                if filter_criteria.get("exclude_bots"):
                    conditions.append("is_bot = 0")

                if filter_criteria.get("name_contains"):
                    conditions.append("(first_name LIKE ? OR last_name LIKE ?)")
                    search = f"%{filter_criteria['name_contains']}%"
                    params.extend([search, search])

            where_clause = " AND ".join(conditions) if conditions else "1=1"
            query = f"""
                SELECT
                    user_id, username, first_name, last_name, phone,
                    bio, is_bot, is_verified, is_premium, has_photo,
                    language_code, scraped_at, last_seen, message_count
                FROM members
                WHERE {where_clause}
                ORDER BY scraped_at DESC
            """

            if filter_criteria and filter_criteria.get("limit"):
                query += f" LIMIT {filter_criteria['limit']}"

            cursor.execute(query, params)
            rows = cursor.fetchall()

            # Convert to list of dicts
            members = [dict(row) for row in rows]

            conn.close()
            logger.info(f"Fetched {len(members)} REAL members from database")
            return members

        except Exception as e:
            logger.error(f"Failed to fetch real member data: {e}")
            return []

    @staticmethod
    def fetch_real_campaign_data() -> List[Dict]:
        """Fetch ACTUAL campaign data from database."""
        import sqlite3

        try:
            conn = sqlite3.connect("campaigns.db")
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            query = """
                SELECT
                    id, name, status, message_template,
                    total_targets, sent_count, failed_count,
                    created_at, started_at, completed_at
                FROM campaigns
                ORDER BY created_at DESC
            """

            cursor.execute(query)
            rows = cursor.fetchall()

            campaigns = [dict(row) for row in rows]

            conn.close()
            logger.info(f"Fetched {len(campaigns)} REAL campaigns from database")
            return campaigns

        except Exception as e:
            logger.error(f"Failed to fetch real campaign data: {e}")
            return []

    @staticmethod
    def fetch_real_account_data() -> List[Dict]:
        """Fetch ACTUAL account data from database."""
        import sqlite3

        from member_scraper import MemberDatabase

        try:
            member_db = MemberDatabase("members.db")
            conn = sqlite3.connect(member_db.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            query = """
                SELECT
                    phone_number, status, session_file,
                    created_at, last_active, messages_sent,
                    is_warmed_up, warmup_stage, proxy_used
                FROM accounts
                ORDER BY created_at DESC
            """

            cursor.execute(query)
            rows = cursor.fetchall()

            accounts = [dict(row) for row in rows]

            conn.close()
            logger.info(f"Fetched {len(accounts)} REAL accounts from database")
            return accounts

        except Exception as e:
            logger.error(f"Failed to fetch real account data: {e}")
            return []

    @staticmethod
    def export_to_csv(data: List[Dict], file_path: str, options: Dict = None) -> bool:
        """Export data to CSV."""
        try:
            if not data:
                logger.warning("No data to export")
                return False

            safe_path = normalize_export_path(file_path)

            with open(safe_path, "w", newline="", encoding="utf-8") as f:
                # Get all unique keys from all dicts
                fieldnames = set()
                for item in data:
                    fieldnames.update(item.keys())

                fieldnames = sorted(fieldnames)

                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)

            logger.info(f"Exported {len(data)} records to CSV: {safe_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export to CSV: {e}")
            return False

    @staticmethod
    def export_to_json(data: List[Dict], file_path: str, options: Dict = None) -> bool:
        """Export data to JSON."""
        try:
            safe_path = normalize_export_path(file_path)

            with open(safe_path, "w", encoding="utf-8") as f:
                json.dump(
                    {"exported_at": datetime.now().isoformat(), "count": len(data), "data": data},
                    f,
                    indent=2,
                    ensure_ascii=False,
                    default=str,
                )

            logger.info(f"Exported {len(data)} records to JSON: {safe_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export to JSON: {e}")
            return False

    @staticmethod
    def export_to_excel(data: List[Dict], file_path: str, options: Dict = None) -> bool:
        """Export data to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export"

            if not data:
                logger.warning("No data to export")
                return False

            # Get headers
            headers = sorted(set().union(*[item.keys() for item in data]))

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=str(item.get(header, "")))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except (TypeError, AttributeError):
                        pass  # Skip cells with None or non-string values
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            safe_path = normalize_export_path(file_path)

            wb.save(safe_path)
            logger.info(f"Exported {len(data)} records to Excel: {safe_path}")
            return True

        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            return False

    @staticmethod
    def export_to_txt(data: List[Dict], file_path: str, options: Dict = None) -> bool:
        """Export data to plain text."""
        try:
            safe_path = normalize_export_path(file_path)

            with open(safe_path, "w", encoding="utf-8") as f:
                f.write(f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Records: {len(data)}\n")
                f.write("=" * 80 + "\n\n")

                for i, item in enumerate(data, 1):
                    f.write(f"Record {i}:\n")
                    for key, value in item.items():
                        f.write(f"  {key}: {value}\n")
                    f.write("\n")

            logger.info(f"Exported {len(data)} records to TXT: {safe_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export to TXT: {e}")
            return False


def show_export_dialog(data_type: str, data_count: int = 0, parent=None) -> Optional[Dict]:
    """Show export dialog and return config if accepted."""
    dialog = DataExportDialog(data_type, data_count, parent)

    if dialog.exec() == QDialog.DialogCode.Accepted:
        return dialog.get_config()

    return None


def quick_export(data: List[Dict], data_type: str, format: str = "csv") -> Optional[str]:
    """Quick export without dialog."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{data_type}_export_{timestamp}.{format}"

    exporter = DataExporter()

    if format == "csv":
        success = exporter.export_to_csv(data, filename)
    elif format == "json":
        success = exporter.export_to_json(data, filename)
    elif format == "xlsx":
        success = exporter.export_to_excel(data, filename)
    else:
        success = exporter.export_to_txt(data, filename)

    return filename if success else None


def export_members_from_database(
    filter_criteria: Dict = None, format: str = "csv", file_path: str = None
) -> bool:
    """Export members directly from database with REAL data."""
    exporter = DataExporter()

    # Fetch REAL data from database
    logger.info("Fetching REAL member data from database...")
    members = exporter.fetch_real_member_data(filter_criteria)

    if not members:
        logger.warning("No members found in database to export")
        return False

    # Generate filename if not provided
    if not file_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"members_export_{timestamp}.{format}"

    # Export using appropriate method
    logger.info(f"Exporting {len(members)} REAL members to {file_path}...")

    if format == "csv":
        success = exporter.export_to_csv(members, file_path)
    elif format == "json":
        success = exporter.export_to_json(members, file_path)
    elif format == "xlsx":
        success = exporter.export_to_excel(members, file_path)
    else:
        success = exporter.export_to_txt(members, file_path)

    if success:
        logger.info(f"✅ Successfully exported {len(members)} REAL members to {file_path}")
    else:
        logger.error(f"❌ Failed to export members to {file_path}")

    return success


def export_campaigns_from_database(format: str = "csv", file_path: str = None) -> bool:
    """Export campaigns directly from database with REAL data."""
    exporter = DataExporter()

    # Fetch REAL data from database
    logger.info("Fetching REAL campaign data from database...")
    campaigns = exporter.fetch_real_campaign_data()

    if not campaigns:
        logger.warning("No campaigns found in database to export")
        return False

    # Generate filename if not provided
    if not file_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"campaigns_export_{timestamp}.{format}"

    # Export
    logger.info(f"Exporting {len(campaigns)} REAL campaigns to {file_path}...")

    if format == "csv":
        success = exporter.export_to_csv(campaigns, file_path)
    elif format == "json":
        success = exporter.export_to_json(campaigns, file_path)
    elif format == "xlsx":
        success = exporter.export_to_excel(campaigns, file_path)
    else:
        success = exporter.export_to_txt(campaigns, file_path)

    if success:
        logger.info(f"✅ Successfully exported {len(campaigns)} REAL campaigns to {file_path}")
    else:
        logger.error(f"❌ Failed to export campaigns to {file_path}")

    return success


def export_accounts_from_database(format: str = "csv", file_path: str = None) -> bool:
    """Export accounts directly from database with REAL data."""
    exporter = DataExporter()

    # Fetch REAL data from database
    logger.info("Fetching REAL account data from database...")
    accounts = exporter.fetch_real_account_data()

    if not accounts:
        logger.warning("No accounts found in database to export")
        return False

    # Generate filename if not provided
    if not file_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"accounts_export_{timestamp}.{format}"

    # Export
    logger.info(f"Exporting {len(accounts)} REAL accounts to {file_path}...")

    if format == "csv":
        success = exporter.export_to_csv(accounts, file_path)
    elif format == "json":
        success = exporter.export_to_json(accounts, file_path)
    elif format == "xlsx":
        success = exporter.export_to_excel(accounts, file_path)
    else:
        success = exporter.export_to_txt(accounts, file_path)

    if success:
        logger.info(f"✅ Successfully exported {len(accounts)} REAL accounts to {file_path}")
    else:
        logger.error(f"❌ Failed to export accounts to {file_path}")

    return success
