"""
Export Manager - Export members and campaigns to various formats.

Supports:
- CSV export
- JSON export
- Excel export (if openpyxl available)
"""

import csv
import json
import logging
import re
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import sqlite3

logger = logging.getLogger(__name__)

# Check for optional Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def _sanitize_filename(filename: str) -> str:
    """Sanitize filenames to avoid invalid characters or traversal."""
    safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", filename or "export")
    return safe_name or "export"


def normalize_export_path(filepath: str) -> Path:
    """Resolve export path safely within the working directory."""
    base_dir = Path.cwd()
    candidate = Path(filepath)
    sanitized_name = _sanitize_filename(candidate.name)

    parent = candidate.parent if candidate.parent != Path('.') else base_dir
    candidate_path = parent / sanitized_name

    try:
        resolved = candidate_path.resolve()
        resolved.relative_to(base_dir)
        return resolved
    except Exception as exc:
        logger.warning(
            "Invalid export path '%s' (%s); falling back to working directory", filepath, exc
        )
        return base_dir / sanitized_name


class ExportManager:
    """Manages data export operations."""
    
    def __init__(self, member_db=None, campaign_db_path: str = "campaigns.db"):
        """Initialize export manager.
        
        Args:
            member_db: MemberDatabase instance
            campaign_db_path: Path to campaigns database
        """
        self.member_db = member_db
        self.campaign_db_path = campaign_db_path
    
    # ==================== MEMBER EXPORTS ====================
    
    def export_members_to_csv(self, filepath: str, filters: Dict = None) -> int:
        """Export members to CSV file.
        
        Args:
            filepath: Output file path
            filters: Optional filters (channel_id, is_safe_target, min_threat_score, etc.)
            
        Returns:
            Number of exported records
        """
        members = self._get_filtered_members(filters)
        
        if not members:
            return 0
        
        # Define CSV columns
        columns = [
            'user_id', 'username', 'first_name', 'last_name', 'phone',
            'bio', 'channel_id', 'channel_username', 'is_bot', 'is_premium',
            'is_safe_target', 'threat_score', 'threat_indicators', 'threat_category',
            'activity_score', 'engagement_potential', 'last_seen', 'scraped_at'
        ]
        
        safe_path = normalize_export_path(filepath)

        with open(safe_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            
            for member in members:
                # Convert to dict and handle special fields
                row = {col: member.get(col, '') for col in columns}
                # Handle list fields
                if isinstance(row.get('threat_indicators'), list):
                    row['threat_indicators'] = '; '.join(row['threat_indicators'])
                writer.writerow(row)
        
        logger.info(f"Exported {len(members)} members to {safe_path}")
        return len(members)
    
    def export_members_to_json(self, filepath: str, filters: Dict = None, pretty: bool = True) -> int:
        """Export members to JSON file.
        
        Args:
            filepath: Output file path
            filters: Optional filters
            pretty: Pretty print JSON
            
        Returns:
            Number of exported records
        """
        members = self._get_filtered_members(filters)
        
        if not members:
            return 0
        
        safe_path = normalize_export_path(filepath)

        with open(safe_path, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(members, f, indent=2, default=str, ensure_ascii=False)
            else:
                json.dump(members, f, default=str, ensure_ascii=False)

        logger.info(f"Exported {len(members)} members to {safe_path}")
        return len(members)
    
    def export_members_to_excel(self, filepath: str, filters: Dict = None) -> int:
        """Export members to Excel file.
        
        Args:
            filepath: Output file path
            filters: Optional filters
            
        Returns:
            Number of exported records
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        members = self._get_filtered_members(filters)
        
        if not members:
            return 0
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        
        # Define columns
        columns = [
            'User ID', 'Username', 'First Name', 'Last Name', 'Phone',
            'Bio', 'Channel ID', 'Channel Username', 'Is Bot', 'Is Premium',
            'Safe Target', 'Threat Score', 'Threat Category', 'Activity Score',
            'Engagement Potential', 'Last Seen', 'Scraped At'
        ]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, member in enumerate(members, 2):
            data = [
                member.get('user_id', ''),
                member.get('username', ''),
                member.get('first_name', ''),
                member.get('last_name', ''),
                member.get('phone', ''),
                member.get('bio', ''),
                member.get('channel_id', ''),
                member.get('channel_username', ''),
                'Yes' if member.get('is_bot') else 'No',
                'Yes' if member.get('is_premium') else 'No',
                'Yes' if member.get('is_safe_target') else 'No',
                member.get('threat_score', 0),
                member.get('threat_category', ''),
                member.get('activity_score', 0),
                member.get('engagement_potential', ''),
                member.get('last_seen', ''),
                member.get('scraped_at', '')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError, ValueError) as e:
                    # Handle cases where cell.value might be None or not convertible to string
                    logger.debug(f"Error processing cell value: {e}")
                    continue
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        safe_path = normalize_export_path(filepath)

        wb.save(safe_path)
        logger.info(f"Exported {len(members)} members to {safe_path}")
        return len(members)
    
    def _get_filtered_members(self, filters: Dict = None) -> List[Dict]:
        """Get members with optional filters.
        
        Args:
            filters: Optional filters dict
            
        Returns:
            List of member dictionaries
        """
        if not self.member_db:
            logger.error("Member database not available")
            return []
        
        filters = filters or {}
        
        # Build query
        query = "SELECT * FROM members WHERE 1=1"
        params = []
        
        if filters.get('channel_id'):
            query += " AND channel_id = ?"
            params.append(filters['channel_id'])
        
        if filters.get('is_safe_target') is not None:
            query += " AND is_safe_target = ?"
            params.append(1 if filters['is_safe_target'] else 0)
        
        if filters.get('min_threat_score') is not None:
            query += " AND threat_score >= ?"
            params.append(filters['min_threat_score'])
        
        if filters.get('max_threat_score') is not None:
            query += " AND threat_score <= ?"
            params.append(filters['max_threat_score'])
        
        if filters.get('has_username'):
            query += " AND username IS NOT NULL AND username != ''"
        
        if filters.get('limit'):
            query += f" LIMIT {int(filters['limit'])}"
        
        try:
            conn = self.member_db._get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(query, params)
            members = [dict(row) for row in cursor.fetchall()]
            return members
        except Exception as e:
            logger.error(f"Error fetching members: {e}")
            return []
    
    # ==================== CAMPAIGN EXPORTS ====================
    
    def export_campaigns_to_csv(self, filepath: str, include_messages: bool = False) -> int:
        """Export campaigns to CSV file.
        
        Args:
            filepath: Output file path
            include_messages: Include individual messages
            
        Returns:
            Number of exported campaigns
        """
        campaigns = self._get_campaigns()
        
        if not campaigns:
            return 0
        
        columns = [
            'id', 'name', 'template', 'status', 'target_channel_id',
            'total_targets', 'sent_count', 'failed_count', 'blocked_count',
            'rate_limit_delay', 'max_messages_per_hour', 'max_messages_per_account',
            'created_at', 'started_at', 'completed_at',
            'scheduled_start', 'scheduled_end', 'active_hours', 'timezone',
            'recurring', 'recurrence_interval'
        ]
        
        safe_path = normalize_export_path(filepath)

        with open(safe_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            
            for campaign in campaigns:
                row = {col: campaign.get(col, '') for col in columns}
                # Format active hours
                if campaign.get('active_hours_start') is not None:
                    row['active_hours'] = f"{campaign.get('active_hours_start', '')}-{campaign.get('active_hours_end', '')}"
                writer.writerow(row)
        
        # Export messages separately if requested
        if include_messages:
            messages_filepath = safe_path.with_suffix("_messages.csv")
            self._export_campaign_messages(messages_filepath)

        logger.info(f"Exported {len(campaigns)} campaigns to {safe_path}")
        return len(campaigns)
    
    def export_campaigns_to_json(self, filepath: str, include_messages: bool = False, pretty: bool = True) -> int:
        """Export campaigns to JSON file.
        
        Args:
            filepath: Output file path
            include_messages: Include individual messages
            pretty: Pretty print JSON
            
        Returns:
            Number of exported campaigns
        """
        campaigns = self._get_campaigns()
        
        if not campaigns:
            return 0
        
        if include_messages:
            for campaign in campaigns:
                campaign['messages'] = self._get_campaign_messages(campaign['id'])
        
        safe_path = normalize_export_path(filepath)

        with open(safe_path, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(campaigns, f, indent=2, default=str, ensure_ascii=False)
            else:
                json.dump(campaigns, f, default=str, ensure_ascii=False)

        logger.info(f"Exported {len(campaigns)} campaigns to {safe_path}")
        return len(campaigns)
    
    def export_campaign_analytics(self, campaign_id: int, filepath: str) -> bool:
        """Export detailed analytics for a specific campaign.
        
        Args:
            campaign_id: Campaign ID
            filepath: Output file path
            
        Returns:
            Success status
        """
        try:
            with sqlite3.connect(self.campaign_db_path) as conn:
                conn.row_factory = sqlite3.Row
                
                # Get campaign details
                campaign = conn.execute(
                    'SELECT * FROM campaigns WHERE id = ?', (campaign_id,)
                ).fetchone()
                
                if not campaign:
                    return False
                
                # Get message statistics
                messages = conn.execute('''
                    SELECT status, COUNT(*) as count 
                    FROM campaign_messages 
                    WHERE campaign_id = ?
                    GROUP BY status
                ''', (campaign_id,)).fetchall()
                
                # Get hourly distribution
                hourly = conn.execute('''
                    SELECT strftime('%H', sent_at) as hour, COUNT(*) as count
                    FROM campaign_messages
                    WHERE campaign_id = ? AND sent_at IS NOT NULL
                    GROUP BY hour
                    ORDER BY hour
                ''', (campaign_id,)).fetchall()
                
                # Get account performance
                accounts = conn.execute('''
                    SELECT account_phone, 
                           COUNT(*) as total,
                           SUM(CASE WHEN status = 'sent' THEN 1 ELSE 0 END) as sent,
                           SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed
                    FROM campaign_messages
                    WHERE campaign_id = ?
                    GROUP BY account_phone
                ''', (campaign_id,)).fetchall()
                
                analytics = {
                    'campaign': dict(campaign),
                    'message_stats': {row['status']: row['count'] for row in messages},
                    'hourly_distribution': {row['hour']: row['count'] for row in hourly},
                    'account_performance': [dict(row) for row in accounts],
                    'exported_at': datetime.now().isoformat()
                }
                
                safe_path = normalize_export_path(filepath)

                with open(safe_path, 'w', encoding='utf-8') as f:
                    json.dump(analytics, f, indent=2, default=str, ensure_ascii=False)

                logger.info(f"Exported analytics for campaign {campaign_id} to {safe_path}")
                return True
                
        except Exception as e:
            logger.error(f"Error exporting campaign analytics: {e}")
            return False
    
    def _get_campaigns(self) -> List[Dict]:
        """Get all campaigns from database."""
        try:
            with sqlite3.connect(self.campaign_db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute('SELECT * FROM campaigns ORDER BY created_at DESC')
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error fetching campaigns: {e}")
            return []
    
    def _get_campaign_messages(self, campaign_id: int) -> List[Dict]:
        """Get all messages for a campaign."""
        try:
            with sqlite3.connect(self.campaign_db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute(
                    'SELECT * FROM campaign_messages WHERE campaign_id = ?',
                    (campaign_id,)
                )
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error fetching campaign messages: {e}")
            return []
    
    def _export_campaign_messages(self, filepath: Path):
        """Export all campaign messages to CSV."""
        try:
            with sqlite3.connect(self.campaign_db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute('SELECT * FROM campaign_messages')
                messages = [dict(row) for row in cursor.fetchall()]
            
            if not messages:
                return
            
            columns = ['id', 'campaign_id', 'user_id', 'account_phone', 
                      'message_text', 'status', 'sent_at', 'error_message', 'retry_count']
            
            safe_path = normalize_export_path(str(filepath))

            with open(safe_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                for msg in messages:
                    row = {col: msg.get(col, '') for col in columns}
                    writer.writerow(row)

            logger.info(f"Exported {len(messages)} messages to {safe_path}")
        except Exception as e:
            logger.error(f"Error exporting messages: {e}")


# ==================== IMPORT FUNCTIONALITY ====================

class ImportManager:
    """Manages data import operations."""
    
    def __init__(self, member_db=None):
        """Initialize import manager.
        
        Args:
            member_db: MemberDatabase instance
        """
        self.member_db = member_db
    
    def import_members_from_csv(self, filepath: str) -> int:
        """Import members from CSV file.
        
        Args:
            filepath: Input CSV file path
            
        Returns:
            Number of imported records
        """
        if not self.member_db:
            logger.error("Member database not available")
            return 0
        
        imported = 0
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                for row in reader:
                    try:
                        member_data = {
                            'user_id': int(row.get('user_id', 0)),
                            'username': row.get('username'),
                            'first_name': row.get('first_name'),
                            'last_name': row.get('last_name'),
                            'phone': row.get('phone'),
                            'bio': row.get('bio'),
                            'channel_id': row.get('channel_id'),
                            'channel_username': row.get('channel_username'),
                            'is_bot': row.get('is_bot', '').lower() in ('true', 'yes', '1'),
                            'is_premium': row.get('is_premium', '').lower() in ('true', 'yes', '1'),
                        }
                        
                        if member_data['user_id']:
                            self.member_db.add_member(member_data)
                            imported += 1
                    except Exception as e:
                        logger.warning(f"Error importing row: {e}")
            
            logger.info(f"Imported {imported} members from {filepath}")
            return imported
            
        except Exception as e:
            logger.error(f"Error importing from CSV: {e}")
            return 0
    
    def import_members_from_json(self, filepath: str) -> int:
        """Import members from JSON file.
        
        Args:
            filepath: Input JSON file path
            
        Returns:
            Number of imported records
        """
        if not self.member_db:
            logger.error("Member database not available")
            return 0
        
        imported = 0
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                members = json.load(f)
            
            if not isinstance(members, list):
                members = [members]
            
            for member_data in members:
                try:
                    if member_data.get('user_id'):
                        self.member_db.add_member(member_data)
                        imported += 1
                except Exception as e:
                    logger.warning(f"Error importing member: {e}")
            
            logger.info(f"Imported {imported} members from {filepath}")
            return imported
            
        except Exception as e:
            logger.error(f"Error importing from JSON: {e}")
            return 0


def get_export_manager(member_db=None, campaign_db_path: str = "campaigns.db") -> ExportManager:
    """Get ExportManager instance.
    
    Args:
        member_db: Optional MemberDatabase instance
        campaign_db_path: Path to campaigns database
        
    Returns:
        ExportManager instance
    """
    return ExportManager(member_db, campaign_db_path)


def get_import_manager(member_db=None) -> ImportManager:
    """Get ImportManager instance.
    
    Args:
        member_db: Optional MemberDatabase instance
        
    Returns:
        ImportManager instance
    """
    return ImportManager(member_db)


